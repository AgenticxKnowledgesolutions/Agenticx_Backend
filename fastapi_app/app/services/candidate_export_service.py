import io
import csv
from datetime import datetime
from typing import Optional, List, Tuple, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_, and_
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.candidate_application import CandidateApplication
from app.services.candidate_service import (
    get_effective_candidate_program,
    CandidateService,
    mask_aadhaar,
    decrypt_aadhaar
)


class CandidateExportService:
    HEADERS = [
        "Application No",
        "Full Name",
        "Preferred Name",
        "Email",
        "Phone",
        "WhatsApp",
        "Gender",
        "Date of Birth",
        "Qualification",
        "University",
        "College",
        "Academic CGPA/%",
        "Passing Year",
        "Academic Status",
        "Experience (Years)",
        "Company",
        "Skills",
        "Program Applied",
        "Program Type",
        "Mode of Learning",
        "Application Status",
        "Aadhaar (Masked)",
        "PAN Number",
        "Address",
        "City",
        "District",
        "State",
        "Standard Fee (INR)",
        "Scholarship (INR)",
        "Admission Fee Paid",
        "Total Paid (INR)",
        "Remaining Balance (INR)",
        "Registration Date"
    ]

    @classmethod
    async def export_candidates(
        cls,
        db: AsyncSession,
        scope: str = "filtered",
        candidate_ids: Optional[List[str]] = None,
        status_filter: Optional[str] = None,
        course_filter: Optional[str] = None,
        qualification_filter: Optional[str] = None,
        search_query: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        format_type: str = "xlsx"
    ) -> Tuple[bytes, str, str]:
        """
        Exports candidate records based on scope ('all', 'filtered', 'selected') and format ('xlsx', 'csv').
        Returns (file_bytes, filename, media_type).
        """
        stmt = select(CandidateApplication).where(CandidateApplication.is_deleted == False)

        if scope == "selected" and candidate_ids:
            stmt = stmt.where(CandidateApplication.id.in_(candidate_ids))
        else:
            # Apply scope == 'filtered' or 'all' filters
            conditions = []
            if status_filter:
                conditions.append(CandidateApplication.application_status == status_filter)
            if course_filter:
                from app.models.program import Program
                stmt = stmt.outerjoin(Program, CandidateApplication.program_id == Program.id)
                conditions.append(
                    or_(
                        CandidateApplication.course_applied.ilike(course_filter),
                        and_(
                            CandidateApplication.program_id == Program.id,
                            Program.name.ilike(course_filter)
                        )
                    )
                )
            if qualification_filter:
                conditions.append(CandidateApplication.qualification.ilike(f"%{qualification_filter}%"))
            if start_date:
                conditions.append(CandidateApplication.created_at >= start_date)
            if end_date:
                conditions.append(CandidateApplication.created_at <= end_date)
            if search_query:
                q = f"%{search_query}%"
                conditions.append(
                    or_(
                        CandidateApplication.full_name.ilike(q),
                        CandidateApplication.email.ilike(q),
                        CandidateApplication.phone.ilike(q),
                        CandidateApplication.application_number.ilike(q)
                    )
                )

            if conditions:
                stmt = stmt.where(and_(*conditions))

        res = await db.execute(stmt.order_by(CandidateApplication.created_at.desc()))
        candidates = res.scalars().all()

        rows = []
        for c in candidates:
            try:
                eff_prog = get_effective_candidate_program(c)
            except Exception:
                eff_prog = {
                    "name": getattr(c, "course_applied", "") or "",
                    "type": getattr(c, "program_type", "Course") or "Course"
                }

            try:
                fin = CandidateService.calculate_financials(c)
            except Exception:
                fin = {
                    "standard_course_fee": getattr(c, "standard_course_fee", 0.0) or 0.0,
                    "scholarship_amount": getattr(c, "scholarship_amount", 0.0) or 0.0,
                    "admission_fee_paid": getattr(c, "admission_fee_paid", False),
                    "total_paid": 0.0,
                    "balance_remaining": 0.0
                }

            # Aadhaar masking
            masked_aadhaar = "XXXX XXXX XXXX"
            try:
                masked_aadhaar = mask_aadhaar(decrypt_aadhaar(c.aadhaar_number_encrypted)) or "XXXX XXXX XXXX"
            except Exception:
                pass

            reg_date = c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""
            dob = c.date_of_birth.strftime("%Y-%m-%d") if c.date_of_birth else ""

            row = [
                c.application_number or "",
                c.full_name or "",
                c.preferred_name or "",
                c.email or "",
                c.phone or "",
                c.whatsapp_number or "",
                c.gender or "",
                dob,
                c.qualification or "",
                c.university_name or "",
                c.college_name or "",
                c.academic_cgpa or c.academic_percentage or "",
                c.passing_year or "",
                c.academic_status or "",
                c.experience_years or "",
                c.company_name or "",
                c.skills or "",
                eff_prog.get("name", "") or c.course_applied or "",
                eff_prog.get("type", "Course") or "Course",
                c.mode_of_learning or "",
                c.application_status or "",
                masked_aadhaar,
                c.pan_number or "",
                c.address or "",
                c.city or "",
                c.district or "",
                c.state or "",
                fin.get("standard_course_fee", 0.0),
                fin.get("scholarship_amount", 0.0),
                "Yes" if fin.get("admission_fee_paid") else "No",
                fin.get("total_paid", 0.0),
                fin.get("balance_remaining", 0.0),
                reg_date
            ]
            rows.append(row)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if format_type.lower() == "csv":
            return cls._generate_csv(rows, timestamp)
        else:
            return cls._generate_excel(rows, timestamp)

    @classmethod
    def _generate_csv(cls, rows: List[List[Any]], timestamp: str) -> Tuple[bytes, str, str]:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(cls.HEADERS)
        writer.writerows(rows)

        # UTF-8-SIG includes BOM so Excel opens CSV cleanly
        content = output.getvalue().encode("utf-8-sig")
        filename = f"candidates_export_{timestamp}.csv"
        return content, filename, "text/csv"

    @classmethod
    def _generate_excel(cls, rows: List[List[Any]], timestamp: str) -> Tuple[bytes, str, str]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates Data"

        # Enable grid lines
        ws.views.sheetView[0].showGridLines = True

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        data_font = Font(name="Calibri", size=10, color="0F172A")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        # Write headers
        ws.append(cls.HEADERS)
        ws.row_dimensions[1].height = 28

        for col_num in range(1, len(cls.HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Write data rows
        for row_idx, row_data in enumerate(rows, start=2):
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 20
            for col_num, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        content = file_stream.getvalue()
        filename = f"candidates_export_{timestamp}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return content, filename, media_type
