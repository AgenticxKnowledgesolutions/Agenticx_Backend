import base64
import hashlib
import uuid
import logging
import asyncio

logger = logging.getLogger(__name__)
import openpyxl
import csv
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from sqlalchemy import select, or_, and_, func, delete
from sqlalchemy.orm import selectinload, Session
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException, status, UploadFile

from app.core.config import settings
from app.models.candidate_application import CandidateApplication, CandidateNote, CandidateTimelineEvent, CandidateImportBatch
from app.models.admin_notification import AdminNotification
from app.models.lead import Lead
from app.models.lead_interaction import LeadInteraction
from app.models.lead_timeline import LeadTimelineEvent as LeadTimelineEventModel
from app.services.upload_service import UploadService
from cryptography.fernet import Fernet

# -------------------------------------------------------------
# Aadhaar Cryptography Helpers
# -------------------------------------------------------------
def get_fernet() -> Fernet:
    # Deterministic Fernet key derived from settings.SECRET_KEY
    key = hashlib.sha256(settings.SECRET_KEY.encode()).digest()
    fernet_key = base64.urlsafe_b64encode(key)
    return Fernet(fernet_key)

def encrypt_aadhaar(aadhaar: str) -> Optional[str]:
    if not aadhaar or not aadhaar.strip():
        return None
    f = get_fernet()
    return f.encrypt(aadhaar.strip().encode()).decode()

def decrypt_aadhaar(encrypted_aadhaar: Optional[str]) -> Optional[str]:
    if not encrypted_aadhaar:
        return None
    f = get_fernet()
    try:
        return f.decrypt(encrypted_aadhaar.encode()).decode()
    except Exception:
        return "Decryption Error"

def mask_aadhaar(aadhaar: Optional[str]) -> Optional[str]:
    if not aadhaar:
        return None
    aadhaar_clean = "".join(filter(str.isdigit, aadhaar))
    if len(aadhaar_clean) >= 4:
        return f"XXXX XXXX {aadhaar_clean[-4:]}"
    return "XXXX XXXX XXXX"

# -------------------------------------------------------------
# Supabase Documents Upload Service
# -------------------------------------------------------------
class CandidateUploadService(UploadService):
    def __init__(self):
        super().__init__()
        self.bucket_name = "candidate-documents"

candidate_upload_service = CandidateUploadService()

def normalize_program_name(raw_name: str) -> str:
    """
    Normalizes course/program title formatting and casing.
    Deduplicates lowercase/uppercase variations (e.g. 'full stack' -> 'Full Stack').
    Preserves known tech/domain acronyms (e.g. AI, ML, FDP, MERN, AWS, UI/UX, etc.) and lowercases small connectors (on, in, of, and, with).
    """
    if not raw_name or not isinstance(raw_name, str):
        return ""
    cleaned = " ".join(raw_name.strip().split())
    if not cleaned:
        return ""

    if not cleaned.islower() and not cleaned.isupper():
        return cleaned

    words = cleaned.split(" ")
    normalized_words = []
    acronyms = {"AI", "ML", "FDP", "MERN", "AWS", "UI/UX", "IT", "HR", "SQL", "API", "REST", "MEAN", "PHP", "SEO", "JS", "CSS", "HTML"}
    small_words = {"on", "in", "of", "and", "with", "for", "to", "at", "by"}

    for idx, w in enumerate(words):
        w_upper = w.upper()
        w_lower = w.lower()
        if w_upper in acronyms:
            normalized_words.append(w_upper)
        elif idx > 0 and w_lower in small_words:
            normalized_words.append(w_lower)
        else:
            normalized_words.append(w.capitalize())

    return " ".join(normalized_words)

def get_effective_candidate_program(candidate: Any) -> dict:
    """
    Determines candidate's effective program name and type safely.
    Handles legacy, custom, and relationship-based course/program data.
    Works for objects (SQLAlchemy model, Row) and dictionaries without triggering async lazy loading I/O.
    """
    program_name = None
    program_type = None
    
    # Safe getter helper to extract column value or dict key without triggering ORM relationship lazy loading
    def get_val(obj, key, default=None):
        if obj is None:
            return default
        if isinstance(obj, dict):
            return obj.get(key, default)
        # For SQLAlchemy models, check __dict__ first to ensure attribute is loaded into instance state
        if hasattr(obj, "__dict__") and key in obj.__dict__:
            return obj.__dict__[key]
        # For standard columns or plain object attributes
        try:
            obj_type = type(obj)
            if hasattr(obj_type, key):
                from sqlalchemy.orm.attributes import InstrumentedAttribute
                if isinstance(getattr(obj_type, key), InstrumentedAttribute):
                    return default
            return getattr(obj, key, default)
        except Exception:
            return default

    # 1. Start with values from linked program relationship ONLY if already loaded into instance state (__dict__)
    prog = None
    if isinstance(candidate, dict):
        prog = candidate.get("program")
    elif hasattr(candidate, "__dict__") and "program" in candidate.__dict__:
        prog = candidate.__dict__["program"]
        
    if prog:
        program_name = get_val(prog, "name")
        program_type = get_val(prog, "program_type")
        
    # 2. Fall back to direct CandidateApplication database fields
    if not program_name:
        program_name = get_val(candidate, "course_applied")
    if not program_type:
        program_type = get_val(candidate, "program_type")
        
    # 3. Check legacy/older fields if any exist on the object
    if not program_name:
        for field in ["course_name", "custom_course_name", "program", "internship", "webinar"]:
            val = get_val(candidate, field)
            if val and isinstance(val, str):
                program_name = val
                if not program_type and field in ["internship", "webinar"]:
                    program_type = field.title()
                break
                
    if program_name and isinstance(program_name, str):
        program_name = normalize_program_name(program_name)
    else:
        program_name = ""
        
    # 4. Infer program type if name is present but type is not
    if program_name and not program_type:
        name_lower = program_name.lower()
        if "webinar" in name_lower:
            program_type = "Webinar"
        elif "fdp" in name_lower or "faculty development" in name_lower:
            program_type = "FDP"
        elif "workshop" in name_lower or "bootcamp" in name_lower:
            program_type = "Workshop"
        elif "internship" in name_lower or "intern" in name_lower:
            program_type = "Internship"
        else:
            program_type = "Course"
            
    if not program_type or not isinstance(program_type, str):
        program_type = "Course"
        
    return {
        "name": program_name,
        "type": program_type
    }

# -------------------------------------------------------------
# Service Core Class
# -------------------------------------------------------------
class CandidateService:
    @staticmethod
    def calculate_document_status(candidate: CandidateApplication) -> str:
        """Calculates status based on completed document fields."""
        fields = [candidate.cv_url, candidate.photo_url, candidate.aadhaar_url]
        filled_count = sum(1 for f in fields if f)
        if filled_count == len(fields):
            return "Complete"
        elif filled_count > 0:
            return "Partial"
        return "Missing Documents"

    @staticmethod
    async def generate_application_number(db: AsyncSession) -> str:
        """Generates a unique application number formatted like CAF-YYYY-XXXXX."""
        current_year = datetime.utcnow().year
        prefix = f"CAF-{current_year}-"
        
        # Select all application numbers for the current year to find the maximum suffix
        stmt = select(CandidateApplication.application_number).where(
            CandidateApplication.application_number.like(f"{prefix}%")
        )
        res = await db.execute(stmt)
        app_numbers = res.scalars().all()
        
        max_seq = 0
        for app_num in app_numbers:
            try:
                # Extract suffix after prefix
                suffix_str = app_num[len(prefix):]
                seq = int(suffix_str)
                if seq > max_seq:
                    max_seq = seq
            except ValueError:
                continue
                
        next_seq = max_seq + 1
        return f"{prefix}{next_seq:05d}"

    @classmethod
    async def create_candidate_application(
        cls, db: AsyncSession, data: Dict[str, Any], created_by: Optional[str] = "Website"
    ) -> CandidateApplication:
        # Check if email or phone already exists in candidate table to prevent duplicates
        email = data.get("email", "").strip().lower()
        phone = data.get("phone", "").strip()
        
        stmt = select(CandidateApplication).where(
            or_(
                CandidateApplication.email.ilike(email),
                CandidateApplication.phone == phone
            ),
            CandidateApplication.is_deleted == False
        )
        existing_res = await db.execute(stmt)
        existing = existing_res.scalars().first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Candidate with email/phone already registered (App Ref: {existing.application_number})"
            )

        app_num = await cls.generate_application_number(db)
        
        # Resolve Program details
        program_id = data.get("program_id")
        course_applied = data.get("course_applied")
        program_type = data.get("program_type")
        mode_of_learning = data.get("mode_of_learning")
        course_duration = data.get("course_duration")
        standard_course_fee = data.get("standard_course_fee", 0.0) or 0.0

        lead_id = data.get("lead_id")
        if not program_id and lead_id:
            lead_stmt = select(Lead).where(Lead.id == lead_id)
            lead_res = await db.execute(lead_stmt)
            lead = lead_res.scalars().first()
            if lead and lead.program_id:
                program_id = lead.program_id

        from app.models.program import Program
        if program_id:
            result_p = await db.execute(select(Program).where(Program.id == program_id))
            program = result_p.scalar_one_or_none()
            if program:
                course_applied = program.name
                program_type = program.program_type
                mode_of_learning = program.mode or mode_of_learning
                course_duration = program.duration or course_duration
                if program.standard_fee and float(program.standard_fee) > 0.0:
                    standard_course_fee = float(program.standard_fee)
        elif course_applied:
            result_p = await db.execute(select(Program).where(Program.name == course_applied, Program.is_deleted == False))
            program = result_p.scalar_one_or_none()
            if program:
                program_id = program.id
                program_type = program.program_type
                mode_of_learning = program.mode or mode_of_learning
                course_duration = program.duration or course_duration
                if program.standard_fee and float(program.standard_fee) > 0.0:
                    standard_course_fee = float(program.standard_fee)

        # Aadhaar Encryption
        aadhaar_plain = data.get("aadhaar_number")
        encrypted_aadhaar = encrypt_aadhaar(aadhaar_plain) if aadhaar_plain else None
        
        # Extract fields
        candidate = CandidateApplication(
            id=str(uuid.uuid4()),
            lead_id=lead_id,
            program_id=program_id,
            application_number=app_num,
            full_name=data.get("full_name"),
            email=email,
            phone=phone,
            whatsapp_number=data.get("whatsapp_number"),
            address=data.get("address"),
            emergency_contact=data.get("emergency_contact"),
            qualification=data.get("qualification"),
            blood_group=data.get("blood_group"),
            course_applied=course_applied,
            program_type=program_type,
            programme_domain=data.get("programme_domain"),
            mode_of_learning=mode_of_learning,
            course_duration=course_duration,
            standard_course_fee=standard_course_fee,
            final_payable_amount=standard_course_fee,
            college_name=data.get("college_name"),
            date_of_birth=data.get("date_of_birth"),
            gender=data.get("gender"),
            reference_details=data.get("reference_details"),
            languages_known=data.get("languages_known"),
            parent_guardian_name=data.get("parent_guardian_name"),
            parent_guardian_occupation=data.get("parent_guardian_occupation"),
            aadhaar_number_encrypted=encrypted_aadhaar,
            registration_transaction_id=data.get("registration_transaction_id"),
            application_status=data.get("application_status", "Submitted"),
            document_status=data.get("document_status", "Missing Documents"),
            candidate_source=data.get("candidate_source", "Website"),
            import_tag=data.get("import_tag"),
            next_followup_at=data.get("next_followup_at"),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        
        # Calculate document status
        candidate.document_status = cls.calculate_document_status(candidate)

        db.add(candidate)

        if data.get("remarks"):
            await cls.add_candidate_note(
                db,
                candidate.id,
                f"[Admission remarks]: {data.get('remarks')}",
                created_by=created_by
            )
        
        # Update associated lead to Converted status if lead_id is provided
        lead_id = data.get("lead_id")
        if lead_id:
            from app.services.lead_service import add_timeline_event as add_lead_timeline_event
            lead_stmt = select(Lead).where(Lead.id == lead_id)
            lead_res = await db.execute(lead_stmt)
            lead = lead_res.scalars().first()
            if lead and lead.status != "Converted":
                lead.status = "Converted"
                await add_lead_timeline_event(
                    db,
                    lead_id=lead_id,
                    event_type="Converted",
                    description="Lead converted to candidate via admission form submission.",
                    created_by="System"
                )
                
        await db.flush()

        # Log timeline event
        await cls.add_timeline_event(
            db,
            candidate.id,
            "Submitted",
            f"Candidate application submitted successfully by {created_by}.",
            created_by
        )

        # Trigger internal admin notification
        notification = AdminNotification(
            id=str(uuid.uuid4()),
            title="New Candidate Application",
            message=f"New application {app_num} received from {candidate.full_name} for course {candidate.course_applied}.",
            notification_type="new_application",
            is_read=False,
            created_at=datetime.utcnow()
        )
        db.add(notification)
        await db.commit()
        await db.refresh(candidate)
        return candidate

    @classmethod
    async def update_existing_candidate_application(
        cls, db: AsyncSession, candidate: CandidateApplication, data: Dict[str, Any], created_by: Optional[str] = "Website"
    ) -> CandidateApplication:
        # Resolve Program details
        program_id = data.get("program_id")
        lead_id = candidate.lead_id or data.get("lead_id")
        if not program_id and lead_id:
            lead_stmt = select(Lead).where(Lead.id == lead_id)
            lead_res = await db.execute(lead_stmt)
            lead = lead_res.scalars().first()
            if lead and lead.program_id:
                program_id = lead.program_id

        from app.models.program import Program
        if program_id:
            candidate.program_id = program_id
            result_p = await db.execute(select(Program).where(Program.id == program_id))
            program = result_p.scalar_one_or_none()
            if program:
                candidate.course_applied = program.name
                candidate.program_type = program.program_type
                candidate.mode_of_learning = program.mode or candidate.mode_of_learning
                candidate.course_duration = program.duration or candidate.course_duration
                
                if program.standard_fee and float(program.standard_fee) > 0.0:
                    candidate.standard_course_fee = float(program.standard_fee)
                elif not candidate.standard_course_fee:
                    candidate.standard_course_fee = 0.0
                
                candidate.final_payable_amount = max(
                    0.0,
                    float(candidate.standard_course_fee) - (
                        (candidate.scholarship_amount or 0.0) +
                        (candidate.special_discount or 0.0) +
                        (candidate.corporate_discount or 0.0) +
                        (candidate.promo_discount or 0.0)
                    )
                )

        # Aadhaar Encryption
        aadhaar_plain = data.get("aadhaar_number")
        if aadhaar_plain:
            candidate.aadhaar_number_encrypted = encrypt_aadhaar(aadhaar_plain)
            
        # Update other fields if they are in data and not None
        for key in [
            "full_name", "email", "phone", "whatsapp_number", "address", 
            "emergency_contact", "qualification", "blood_group", "course_applied", 
            "mode_of_learning", "college_name", "date_of_birth", "gender", 
            "reference_details", "languages_known", "parent_guardian_name", 
            "parent_guardian_occupation", "registration_transaction_id"
        ]:
            if key in data and data[key] is not None:
                val = data[key]
                if isinstance(val, str):
                    val = val.strip()
                if key == "email" and val:
                    val = val.lower()
                setattr(candidate, key, val)

        candidate.application_status = data.get("application_status", "Submitted")
        candidate.updated_at = datetime.utcnow()
        candidate.document_status = cls.calculate_document_status(candidate)

        if data.get("remarks"):
            await cls.add_candidate_note(
                db,
                candidate.id,
                f"[Admission remarks]: {data.get('remarks')}",
                created_by=created_by
            )
        
        # Update associated lead to Converted status if lead_id is provided
        lead_id = candidate.lead_id or data.get("lead_id")
        if lead_id:
            from app.services.lead_service import add_timeline_event as add_lead_timeline_event
            lead_stmt = select(Lead).where(Lead.id == lead_id)
            lead_res = await db.execute(lead_stmt)
            lead = lead_res.scalars().first()
            if lead and lead.status != "Converted":
                lead.status = "Converted"
                await add_lead_timeline_event(
                    db,
                    lead_id=lead_id,
                    event_type="Converted",
                    description="Lead converted to candidate via admission form submission.",
                    created_by="System"
                )
                
        await db.flush()

        # Log timeline event
        await cls.add_timeline_event(
            db,
            candidate.id,
            "Submitted",
            f"Candidate application updated/completed successfully by {created_by}.",
            created_by
        )

        # Trigger internal admin notification
        notification = AdminNotification(
            id=str(uuid.uuid4()),
            title="Candidate Application Completed",
            message=f"Application {candidate.application_number} completed by {candidate.full_name} for course {candidate.course_applied}.",
            notification_type="new_application",
            is_read=False,
            created_at=datetime.utcnow()
        )
        db.add(notification)
        await db.commit()
        await db.refresh(candidate)
        return candidate

    @staticmethod
    async def add_timeline_event(
        db: AsyncSession, candidate_id: str, event_type: str, description: str, created_by: Optional[str] = "Admin"
    ) -> Optional[CandidateTimelineEvent]:
        # Disabled: Timeline Events feature removed
        return None

    @staticmethod
    async def add_candidate_note(
        db: AsyncSession, candidate_id: str, content: str, created_by: Optional[str] = "Admin"
    ) -> CandidateNote:
        note = CandidateNote(
            id=str(uuid.uuid4()),
            candidate_id=candidate_id,
            content=content,
            created_by=created_by,
            created_at=datetime.utcnow()
        )
        db.add(note)
        await db.flush()
        return note

    @classmethod
    async def upload_document(
        cls, db: AsyncSession, candidate_id: str, doc_type: str, file_bytes: bytes, original_filename: str, mime_type: str, user_email: Optional[str] = "Admin"
    ) -> CandidateApplication:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        # Map document type to fields
        field_mapping = {
            "cv": "cv_url",
            "photo": "photo_url",
            "aadhaar": "aadhaar_url",
            "college-id": "college_id_url",
            "confirmation-letter": "confirmation_letter_url"
        }
        
        field_name = field_mapping.get(doc_type)
        if not field_name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid document type")

        # Upload file to Supabase in structured folder
        folder = f"{doc_type}/{candidate.id}"
        public_url = await candidate_upload_service.upload_file(file_content=file_bytes, folder=folder, original_filename=original_filename, mime_type=mime_type)

        # Update candidate URL field
        setattr(candidate, field_name, public_url)
        candidate.document_status = cls.calculate_document_status(candidate)
        candidate.updated_at = datetime.utcnow()
        
        # Add timeline event
        await cls.add_timeline_event(
            db,
            candidate.id,
            "Document Uploaded",
            f"Uploaded {doc_type.replace('-', ' ').title()} document: {original_filename}",
            user_email
        )
        await db.commit()
        await db.refresh(candidate)
        return candidate

    @classmethod
    async def get_applications(
        cls,
        db: AsyncSession,
        status_filter: Optional[str] = None,
        course_filter: Optional[str] = None,
        qualification_filter: Optional[str] = None,
        search_query: Optional[str] = None,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        skip: int = 0,
        limit: int = 50,
        is_deleted: bool = False
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication)
        count_stmt = select(func.count(CandidateApplication.id))
        
        conditions = [CandidateApplication.is_deleted == is_deleted]
        if status_filter:
            conditions.append(CandidateApplication.application_status == status_filter)
        if course_filter:
            from app.models.program import Program
            stmt = stmt.outerjoin(Program, CandidateApplication.program_id == Program.id)
            count_stmt = count_stmt.outerjoin(Program, CandidateApplication.program_id == Program.id)
            conditions.append(
                or_(
                    CandidateApplication.course_applied.ilike(course_filter),
                    and_(
                        CandidateApplication.program_id == Program.id,
                        Program.name.ilike(course_filter)
                    )
                )
            )
        if qualification_filter:
            conditions.append(CandidateApplication.qualification.ilike(f"%{qualification_filter}%"))
        if start_date:
            conditions.append(CandidateApplication.created_at >= start_date)
        if end_date:
            conditions.append(CandidateApplication.created_at <= end_date)
            
        if search_query:
            q = f"%{search_query}%"
            conditions.append(
                or_(
                    CandidateApplication.full_name.ilike(q),
                    CandidateApplication.email.ilike(q),
                    CandidateApplication.phone.ilike(q),
                    CandidateApplication.application_number.ilike(q)
                )
            )

        if conditions:
            stmt = stmt.where(and_(*conditions))
            count_stmt = count_stmt.where(and_(*conditions))

        # Order by created_at desc
        stmt = stmt.order_by(CandidateApplication.created_at.desc()).offset(skip).limit(limit)
        
        # Execute queries
        records_res = await db.execute(stmt)
        records = records_res.scalars().all()
        
        total_res = await db.execute(count_stmt)
        total = total_res.scalar() or 0

        # Format output (mask Aadhaar for all records in the list view)
        formatted_records = []
        for r in records:
            r_dict = {c.name: getattr(r, c.name) for c in r.__table__.columns}
            try:
                r_dict["aadhaar_number_masked"] = mask_aadhaar(decrypt_aadhaar(r.aadhaar_number_encrypted))
            except Exception:
                r_dict["aadhaar_number_masked"] = "XXXX XXXX XXXX"

            try:
                eff_prog = get_effective_candidate_program(r)
                r_dict["effective_program_name"] = eff_prog.get("name", "")
                r_dict["effective_program_type"] = eff_prog.get("type", "Course")
            except Exception:
                r_dict["effective_program_name"] = getattr(r, "course_applied", "") or ""
                r_dict["effective_program_type"] = getattr(r, "program_type", "Course") or "Course"

            formatted_records.append(r_dict)

        return {"total": total, "records": formatted_records}

    @classmethod
    async def get_program_options(cls, db: AsyncSession) -> List[Dict[str, Any]]:
        """
        Retrieves a dynamic, aggregated list of all programs/courses used by candidates.
        """
        from app.models.program import Program
        stmt = (
            select(
                CandidateApplication.course_applied,
                CandidateApplication.program_type,
                Program.name.label("prog_name"),
                Program.program_type.label("prog_type"),
                func.count(CandidateApplication.id).label("count")
            )
            .outerjoin(Program, CandidateApplication.program_id == Program.id)
            .where(
                CandidateApplication.is_deleted == False
            )
            .group_by(
                CandidateApplication.course_applied,
                CandidateApplication.program_type,
                Program.name,
                Program.program_type
            )
        )
        res = await db.execute(stmt)
        rows = res.all()

        aggregated = {}
        for course_applied, program_type, prog_name, prog_type, count in rows:
            class MockProgram:
                def __init__(self, name, ptype):
                    self.name = name
                    self.program_type = ptype

            class MockCandidate:
                def __init__(self, ca, pt, prog):
                    self.course_applied = ca
                    self.program_type = pt
                    self.program = prog

            prog_obj = MockProgram(prog_name, prog_type) if prog_name else None
            mock_cand = MockCandidate(course_applied, program_type, prog_obj)
            resolved = get_effective_candidate_program(mock_cand)
            name = resolved["name"]
            type_val = resolved["type"]

            if not name:
                continue

            key = (name.lower(), type_val.lower())
            if key not in aggregated:
                aggregated[key] = {
                    "name": name,
                    "type": type_val,
                    "count": 0
                }
            aggregated[key]["count"] += count

        results = list(aggregated.values())
        results.sort(key=lambda x: x["name"].lower())
        return results

    @classmethod
    def calculate_financials(
        cls,
        candidate: Any,
        current_payment: Optional[Any] = None
    ) -> Dict[str, Any]:
        """
        Authoritative centralized financial calculation helper.

        Business Rules:
        1. final_payable_amount represents the total Course Fee.
        2. admission_fee_amount is mandatory and independent of Course Fee.
        3. Admission Fee NEVER reduces Course Fee, Final Payable Amount, or Remaining Course Balance.
        4. course_paid is the sum of all Paid payments where payment_type != 'Admission Fee'.
        5. admission_fee_paid_amount is the sum of all Paid payments where payment_type == 'Admission Fee'.
        6. remaining_course_balance = max(0.0, final_payable_amount - course_paid).
        7. total_collected = admission_fee_paid_amount + course_paid.
        """
        payments = []
        if isinstance(candidate, dict):
            payments = list(candidate.get("payments") or [])
        elif hasattr(candidate, "__dict__") and "payments" in candidate.__dict__:
            payments = list(candidate.__dict__["payments"] or [])

        if current_payment and current_payment.status == "Paid":
            if not any(p.id == current_payment.id for p in payments):
                payments.append(current_payment)

        admission_fee_paid_amount = sum(
            float(p.amount) for p in payments
            if p.status == "Paid" and p.payment_type == "Admission Fee"
        )

        course_paid_amount = sum(
            float(p.amount) for p in payments
            if p.status == "Paid" and p.payment_type != "Admission Fee"
        )

        final_payable = float(getattr(candidate, "final_payable_amount", 0.0) or 0.0)
        admission_fee_req = float(getattr(candidate, "admission_fee_amount", 0.0) or 0.0)

        is_admission_paid = bool(getattr(candidate, "admission_fee_paid", False)) or (admission_fee_paid_amount > 0)

        remaining_course_balance = max(0.0, final_payable - course_paid_amount)
        total_collected = admission_fee_paid_amount + course_paid_amount

        return {
            "final_payable_amount": round(final_payable, 2),
            "admission_fee_amount": round(admission_fee_req, 2),
            "admission_fee_paid": is_admission_paid,
            "admission_fee_paid_amount": round(admission_fee_paid_amount, 2),
            "course_paid": round(course_paid_amount, 2),
            "remaining_course_balance": round(remaining_course_balance, 2),
            "total_collected": round(total_collected, 2)
        }

    @classmethod
    async def get_application_by_id(cls, db: AsyncSession, candidate_id: str) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id).options(
            selectinload(CandidateApplication.notes),
            selectinload(CandidateApplication.timeline_events),
            selectinload(CandidateApplication.payments),
            selectinload(CandidateApplication.program)
        )
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate application not found")

        r_dict = {c.name: getattr(candidate, c.name) for c in candidate.__table__.columns}
        decrypted = decrypt_aadhaar(candidate.aadhaar_number_encrypted)
        r_dict["aadhaar_number_decrypted"] = decrypted
        r_dict["aadhaar_number_masked"] = mask_aadhaar(decrypted)

        # Attach centralized financials
        financials = cls.calculate_financials(candidate)
        r_dict["financials"] = financials
        r_dict["course_paid"] = financials["course_paid"]
        r_dict["remaining_balance"] = financials["remaining_course_balance"]
        r_dict["total_collected"] = financials["total_collected"]

        # Serialize notes, timeline events, and payments
        r_dict["notes"] = [
            {"id": n.id, "content": n.content, "created_by": n.created_by, "created_at": n.created_at} for n in candidate.notes
        ]
        r_dict["timeline_events"] = [
            {"id": t.id, "event_type": t.event_type, "description": t.description, "created_by": t.created_by, "created_at": t.created_at} for t in candidate.timeline_events
        ]
        r_dict["payments"] = [
            {
                "id": p.id,
                "amount": p.amount,
                "payment_type": p.payment_type,
                "payment_method": p.payment_method,
                "status": p.status,
                "transaction_id": p.transaction_id,
                "payment_date": p.payment_date,
                "created_at": p.created_at,
                "receipt_number": p.receipt_number,
                "receipt_url": p.receipt_url
            } for p in candidate.payments
        ]
        return r_dict

    @classmethod
    async def update_personal_info(
        cls, db: AsyncSession, candidate_id: str, data: Dict[str, Any], user_email: Optional[str] = "Admin"
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        # Aadhaar handling
        if "aadhaar_number" in data and data["aadhaar_number"] is not None:
            plain_aadhaar = str(data["aadhaar_number"]).strip()
            if plain_aadhaar:
                candidate.aadhaar_number_encrypted = encrypt_aadhaar(plain_aadhaar)
            else:
                candidate.aadhaar_number_encrypted = None

        updatable_fields = [
            "full_name", "preferred_name", "gender", "date_of_birth", "email",
            "phone", "whatsapp_number", "pan_number", "address", "city",
            "district", "state", "country", "pincode", "emergency_contact",
            "parent_guardian_name", "parent_guardian_occupation",
            "parent_guardian_phone", "parent_guardian_relationship"
        ]
        updated_keys = []
        for key in updatable_fields:
            if key in data and data[key] is not None:
                setattr(candidate, key, data[key])
                updated_keys.append(key)

        candidate.updated_at = datetime.utcnow()
        await cls.add_timeline_event(
            db, candidate.id, "Personal Info Updated",
            f"Personal information updated by {user_email}. Fields: {', '.join(updated_keys) if updated_keys else 'Aadhaar'}.", user_email
        )
        await db.commit()
        return await cls.get_application_by_id(db, candidate.id)

    @classmethod
    async def update_academic_info(
        cls, db: AsyncSession, candidate_id: str, data: Dict[str, Any], user_email: Optional[str] = "Admin"
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        updatable_fields = [
            "sslc_details", "plus_two_details", "diploma_details", "ug_details", "pg_details",
            "university_name", "college_name", "qualification", "academic_percentage",
            "academic_cgpa", "passing_year", "academic_status"
        ]
        updated_keys = []
        for key in updatable_fields:
            if key in data and data[key] is not None:
                setattr(candidate, key, data[key])
                updated_keys.append(key)

        candidate.updated_at = datetime.utcnow()
        await cls.add_timeline_event(
            db, candidate.id, "Academic Info Updated",
            f"Academic information updated by {user_email}. Fields: {', '.join(updated_keys)}.", user_email
        )
        await db.commit()
        return await cls.get_application_by_id(db, candidate.id)

    @classmethod
    async def update_professional_info(
        cls, db: AsyncSession, candidate_id: str, data: Dict[str, Any], user_email: Optional[str] = "Admin"
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        updatable_fields = [
            "experience_years", "company_name", "skills", "cv_url", "linkedin_url", "portfolio_url"
        ]
        updated_keys = []
        for key in updatable_fields:
            if key in data and data[key] is not None:
                setattr(candidate, key, data[key])
                updated_keys.append(key)

        candidate.updated_at = datetime.utcnow()
        await cls.add_timeline_event(
            db, candidate.id, "Professional Info Updated",
            f"Professional details updated by {user_email}. Fields: {', '.join(updated_keys)}.", user_email
        )
        await db.commit()
        return await cls.get_application_by_id(db, candidate.id)

    @classmethod
    async def update_program_info(
        cls, db: AsyncSession, candidate_id: str, data: Dict[str, Any], user_email: Optional[str] = "Admin"
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        program_id = data.get("program_id")
        if program_id and program_id != candidate.program_id:
            from app.models.program import Program
            prog_res = await db.execute(select(Program).where(Program.id == program_id))
            program = prog_res.scalar_one_or_none()
            if program:
                candidate.program_id = program.id
                candidate.course_applied = program.name
                candidate.program_type = program.program_type
                if program.duration:
                    candidate.course_duration = program.duration
                if program.mode:
                    candidate.mode_of_learning = program.mode

        updatable_fields = [
            "program_type", "course_applied", "batch_name", "trainer_name",
            "course_start_date", "completed_at", "course_duration",
            "mode_of_learning", "training_location", "programme_domain"
        ]
        updated_keys = []
        for key in updatable_fields:
            if key in data and data[key] is not None:
                setattr(candidate, key, data[key])
                updated_keys.append(key)

        candidate.updated_at = datetime.utcnow()
        await cls.add_timeline_event(
            db, candidate.id, "Program Info Updated",
            f"Program details updated by {user_email}. Fields: {', '.join(updated_keys)}.", user_email
        )
        await db.commit()
        return await cls.get_application_by_id(db, candidate.id)

    @classmethod
    async def update_fee_info(
        cls, db: AsyncSession, candidate_id: str, data: Dict[str, Any], user_email: Optional[str] = "Admin"
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        updatable_fields = [
            "standard_course_fee", "scholarship_amount", "special_discount",
            "corporate_discount", "promo_discount", "gst_percentage", "gst_amount",
            "convenience_fee", "admission_fee_amount", "booking_amount",
            "offer_remarks", "offer_expiry_date", "admission_fee_paid", "auto_enroll_enabled"
        ]
        for key in updatable_fields:
            if key in data and data[key] is not None:
                setattr(candidate, key, data[key])

        # Dynamic recalculation of total discounts, GST, and final payable amount
        std_fee = candidate.standard_course_fee or 0.0
        tot_discount = (
            (candidate.scholarship_amount or 0.0) +
            (candidate.special_discount or 0.0) +
            (candidate.corporate_discount or 0.0) +
            (candidate.promo_discount or 0.0)
        )
        taxable = max(0.0, std_fee - tot_discount)
        if (candidate.gst_percentage or 0.0) > 0:
            candidate.gst_amount = round(taxable * (candidate.gst_percentage / 100.0), 2)
        gst = candidate.gst_amount or 0.0
        conv = candidate.convenience_fee or 0.0

        candidate.final_payable_amount = round(taxable + gst + conv, 2)
        candidate.updated_at = datetime.utcnow()

        await cls.add_timeline_event(
            db, candidate.id, "Fee Details Updated",
            f"Financial details updated by {user_email}. Final Payable Amount: ₹{candidate.final_payable_amount}.", user_email
        )
        await db.commit()
        return await cls.get_application_by_id(db, candidate.id)

    @classmethod
    async def update_certificate_info(
        cls, db: AsyncSession, candidate_id: str, data: Dict[str, Any], user_email: Optional[str] = "Admin"
    ) -> Dict[str, Any]:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

        updatable_fields = [
            "certificate_program_type", "certificate_course_name", "certificate_partner",
            "certificate_topics", "certificate_domain", "certificate_duration",
            "certificate_mode", "certificate_body_override", "certificate_title_override",
            "certificate_completion_date", "certificate_issue_date"
        ]
        for key in updatable_fields:
            if key in data and data[key] is not None:
                setattr(candidate, key, data[key])

        candidate.updated_at = datetime.utcnow()
        await cls.add_timeline_event(
            db, candidate.id, "Certificate Metadata Overrides Updated",
            f"Candidate-level certificate metadata overrides updated by {user_email}.", user_email
        )
        await db.commit()
        return await cls.get_application_by_id(db, candidate.id)

    @classmethod
    async def soft_delete_application(
        cls, db: AsyncSession, candidate_id: str, user_email: Optional[str] = "Admin"
    ) -> bool:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == False)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            return False
        
        candidate.is_deleted = True
        candidate.deleted_at = datetime.utcnow()
        
        # Log timeline event
        await cls.add_timeline_event(
            db,
            candidate.id,
            "Deleted",
            f"Candidate application moved to trash by {user_email}.",
            user_email
        )
        await db.commit()
        return True

    @classmethod
    async def bulk_soft_delete_applications(
        cls, db: AsyncSession, candidate_ids: List[str], user_email: Optional[str] = "Admin"
    ) -> int:
        if not candidate_ids:
            return 0
        
        stmt = select(CandidateApplication).where(
            CandidateApplication.id.in_(candidate_ids),
            CandidateApplication.is_deleted == False
        )
        res = await db.execute(stmt)
        candidates = res.scalars().all()
        
        deleted_count = 0
        from app.models.lead_token import LeadToken
        for candidate in candidates:
            candidate.is_deleted = True
            candidate.deleted_at = datetime.utcnow()
            await cls.add_timeline_event(
                db,
                candidate.id,
                "Deleted",
                f"Candidate application moved to trash by bulk operation by {user_email}.",
                user_email
            )
            # Reset associated lead status and delete tokens
            if candidate.lead_id:
                lead_stmt = select(Lead).where(Lead.id == candidate.lead_id)
                lead_res = await db.execute(lead_stmt)
                lead = lead_res.scalar_one_or_none()
                if lead:
                    lead.status = "qualified"
                    await db.execute(delete(LeadToken).where(LeadToken.lead_id == lead.id))
            deleted_count += 1
            
        await db.commit()
        return deleted_count

    @classmethod
    async def restore_application(
        cls, db: AsyncSession, candidate_id: str, user_email: Optional[str] = "Admin"
    ) -> bool:
        stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id, CandidateApplication.is_deleted == True)
        res = await db.execute(stmt)
        candidate = res.scalar_one_or_none()
        if not candidate:
            return False
        
        candidate.is_deleted = False
        candidate.deleted_at = None
        
        # Log timeline event
        await cls.add_timeline_event(
            db,
            candidate.id,
            "Restored",
            f"Candidate application restored from trash by {user_email}.",
            user_email
        )
        await db.commit()
        return True

    @classmethod
    async def hard_delete_application(
        cls, db: AsyncSession, candidate_id: str
    ) -> bool:
        count = await cls.bulk_hard_delete_applications(db, [candidate_id])
        return count > 0

    @classmethod
    async def bulk_hard_delete_applications(
        cls, db: AsyncSession, candidate_ids: List[str]
    ) -> int:
        if not candidate_ids:
            return 0

        # 1. Fetch candidates to get their file URLs
        stmt = select(CandidateApplication).where(CandidateApplication.id.in_(candidate_ids))
        res = await db.execute(stmt)
        candidates = res.scalars().all()
        if not candidates:
            return 0

        # Collect all file deletion tasks
        delete_tasks = []
        
        # Helper for candidate documents
        async def safe_delete_doc(url):
            try:
                await candidate_upload_service.delete_file(url)
            except Exception as e:
                logger.error(f"Error deleting candidate document {url}: {e}")

        # Helper for certificates
        async def safe_delete_cert(url):
            try:
                from app.services.certificate_service import CertificateUploadService
                cert_uploader = CertificateUploadService()
                await cert_uploader.delete_file(url)
            except Exception as e:
                logger.error(f"Error deleting candidate certificate {url}: {e}")

        for candidate in candidates:
            # Candidate documents (candidate-documents bucket)
            doc_urls = [
                candidate.cv_url,
                candidate.photo_url,
                candidate.aadhaar_url,
                candidate.college_id_url,
                candidate.confirmation_letter_url
            ]
            for url in doc_urls:
                if url:
                    delete_tasks.append(safe_delete_doc(url))

            # Certificate (certificates bucket)
            if candidate.certificate_url:
                delete_tasks.append(safe_delete_cert(candidate.certificate_url))

        # Run file deletions in parallel
        if delete_tasks:
            await asyncio.gather(*delete_tasks, return_exceptions=True)

        # 2. Delete candidates from database
        deleted_count = 0
        from app.models.lead_token import LeadToken
        for candidate in candidates:
            # Reset associated lead status and delete tokens
            if candidate.lead_id:
                lead_stmt = select(Lead).where(Lead.id == candidate.lead_id)
                lead_res = await db.execute(lead_stmt)
                lead = lead_res.scalar_one_or_none()
                if lead:
                    lead.status = "qualified"
                    await db.execute(delete(LeadToken).where(LeadToken.lead_id == lead.id))
            await db.delete(candidate)
            deleted_count += 1

        await db.commit()
        return deleted_count

    @classmethod
    async def bulk_regenerate_certificates(
        cls, db: AsyncSession, candidate_ids: List[str]
    ) -> dict:
        results = []
        processed = 0
        success_count = 0
        failed_count = 0

        # Load candidates
        stmt = select(CandidateApplication).where(CandidateApplication.id.in_(candidate_ids))
        res = await db.execute(stmt)
        candidates = res.scalars().all()

        from app.services.certificate_service import certificate_service

        for candidate in candidates:
            processed += 1
            # Match status check case-insensitively
            status_lower = (candidate.application_status or "").lower()
            if status_lower != "completed":
                results.append({
                    "id": candidate.id,
                    "success": False,
                    "error": "Candidate application status is not 'Completed'"
                })
                failed_count += 1
                continue

            try:
                # Regenerate certificate
                await certificate_service.regenerate_certificate(db, candidate)
                # Commit change for each candidate to avoid losing all progress if one fails
                await db.commit()
                results.append({
                    "id": candidate.id,
                    "success": True,
                    "certificate_url": candidate.certificate_url
                })
                success_count += 1
            except Exception as e:
                logger.error(f"Failed to regenerate certificate for candidate {candidate.id}: {e}", exc_info=True)
                results.append({
                    "id": candidate.id,
                    "success": False,
                    "error": str(e)
                })
                failed_count += 1

        return {
            "processed": processed,
            "success_count": success_count,
            "failed_count": failed_count,
            "results": results
        }

    @classmethod
    async def update_application_status(
        cls,
        db: AsyncSession,
        candidate_id: str,
        new_status: str,
        course_start_date: Optional[datetime] = None,
        completed_at: Optional[datetime] = None,
        course_duration: Optional[str] = None,
        performance: Optional[str] = None,
        program_type: Optional[str] = None,
        course_applied: Optional[str] = None,
        user_email: Optional[str] = "Admin",
        program_id: Optional[str] = None,
        programme_domain: Optional[str] = None,
        college_name: Optional[str] = None
    ) -> CandidateApplication:
        try:
            stmt = select(CandidateApplication).where(CandidateApplication.id == candidate_id)
            res = await db.execute(stmt)
            candidate = res.scalar_one_or_none()
            if not candidate:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Candidate not found")

            # Lowercase incoming status
            status_val = new_status.lower()
            
            # Validation before changing status to Completed
            if status_val == "completed":
                perf_val = performance if performance is not None else candidate.performance
                prog_val = program_type if program_type is not None else candidate.program_type
                if not prog_val:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Cannot mark status as Completed: Program Type is required."
                    )
                if prog_val != "Faculty Development Programme" and not perf_val:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Cannot mark status as Completed: Performance is required for student programs."
                    )

            # Check if any certificate fields changed before updating them
            cert_fields_changed = False
            if status_val == "completed" and candidate.certificate_status == "valid":
                if course_start_date is not None and course_start_date != candidate.course_start_date:
                    cert_fields_changed = True
                if completed_at is not None and completed_at != candidate.completed_at:
                    cert_fields_changed = True
                if course_duration is not None and course_duration != candidate.course_duration:
                    cert_fields_changed = True
                if performance is not None and performance != candidate.performance:
                    cert_fields_changed = True
                if program_type is not None and program_type != candidate.program_type:
                    cert_fields_changed = True
                if course_applied is not None and course_applied != candidate.course_applied:
                    cert_fields_changed = True
                if programme_domain is not None and programme_domain != candidate.programme_domain:
                    cert_fields_changed = True

            old_status = candidate.application_status
            candidate.application_status = status_val

            # Resolve Program details first if program_id is passed
            if program_id is not None:
                candidate.program_id = program_id
                if program_id:
                    from app.models.program import Program
                    result_p = await db.execute(select(Program).where(Program.id == program_id))
                    prog = result_p.scalar_one_or_none()
                    if prog:
                        candidate.course_applied = prog.name
                        candidate.program_type = prog.program_type
                        candidate.course_duration = prog.duration
                        
                        if prog.standard_fee and float(prog.standard_fee) > 0.0:
                            candidate.standard_course_fee = float(prog.standard_fee)
                        elif not candidate.standard_course_fee:
                            candidate.standard_course_fee = 0.0
                            
                        candidate.mode_of_learning = prog.mode
                        candidate.final_payable_amount = max(
                            0.0,
                            float(candidate.standard_course_fee) - (
                                (candidate.scholarship_amount or 0.0) +
                                (candidate.special_discount or 0.0) +
                                (candidate.corporate_discount or 0.0) +
                                (candidate.promo_discount or 0.0)
                            )
                        )
            elif course_applied is not None and course_applied != candidate.course_applied:
                from app.models.program import Program
                result_p = await db.execute(select(Program).where(Program.name == course_applied, Program.is_deleted == False))
                prog = result_p.scalar_one_or_none()
                if prog:
                    candidate.program_id = prog.id
                    candidate.program_type = prog.program_type
                    candidate.course_duration = prog.duration
                    
                    if prog.standard_fee and float(prog.standard_fee) > 0.0:
                        candidate.standard_course_fee = float(prog.standard_fee)
                    elif not candidate.standard_course_fee:
                        candidate.standard_course_fee = 0.0
                        
                    candidate.final_payable_amount = max(
                        0.0,
                        float(candidate.standard_course_fee) - (
                            (candidate.scholarship_amount or 0.0) +
                            (candidate.special_discount or 0.0) +
                            (candidate.corporate_discount or 0.0) +
                            (candidate.promo_discount or 0.0)
                        )
                    )

            # Update course start/end dates and duration (Administrative overrides)
            if course_start_date is not None:
                candidate.course_start_date = course_start_date
            if completed_at is not None:
                candidate.completed_at = completed_at
            if course_duration is not None:
                candidate.course_duration = course_duration
            if performance is not None:
                candidate.performance = performance
            if program_type is not None:
                candidate.program_type = program_type
            if course_applied is not None:
                candidate.course_applied = course_applied
            if programme_domain is not None:
                candidate.programme_domain = programme_domain
            if college_name is not None:
                candidate.college_name = college_name

            candidate.updated_at = datetime.utcnow()

            # Add timeline event
            desc = f"Application status changed from '{old_status}' to '{status_val}'."
            await cls.add_timeline_event(db, candidate.id, status_val, desc, user_email)

            # Update DB FIRST and commit status change
            await db.commit()
            await db.refresh(candidate)

            # Trigger certificate generation if status updated to Completed OR certificate fields changed
            if status_val == "completed" and (candidate.certificate_status != "valid" or cert_fields_changed):
                try:
                    # Validate required fields
                    if not candidate.full_name:
                        raise Exception("Missing name")
                    if not candidate.course_applied:
                        raise Exception("Missing course")
                    if not candidate.date_of_birth:
                        raise Exception("Missing DOB")
                    if candidate.program_type != "Faculty Development Programme" and not candidate.performance:
                        raise Exception("Missing performance")
                    if not candidate.program_type:
                        raise Exception("Missing program type")

                    from app.services.certificate_service import certificate_service
                    if cert_fields_changed:
                        await certificate_service.regenerate_certificate(db, candidate)
                    else:
                        await certificate_service.generate_and_save_certificate(db, candidate)
                    await db.commit()
                    await db.refresh(candidate)
                except Exception as ce:
                    print("CERTIFICATE ERROR:", str(ce))
                    # DO NOT break status update

            return candidate
        except HTTPException:
            # Re-raise standard FastAPI HTTPExceptions
            raise
        except Exception as e:
            print("STATUS UPDATE ERROR:", str(e))
            raise HTTPException(status_code=500, detail=str(e))

    # -------------------------------------------------------------
    # Excel/CSV Import Parser Methods
    # -------------------------------------------------------------
    @staticmethod
    def parse_file_headers(file_bytes: bytes, filename: str) -> Dict[str, Any]:
        """Reads file headers and returns the first 20 rows for layout mapping previews."""
        headers = []
        rows = []
        
        if filename.endswith(".csv"):
            try:
                content = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                content = file_bytes.decode("latin-1")
            
            reader = csv.reader(io.StringIO(content))
            for i, r in enumerate(reader):
                if i == 0:
                    headers = [h.strip() for h in r]
                else:
                    rows.append(r)
                    if len(rows) >= 20:
                        break
        else:
            # openpyxl for xlsx files
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = wb.active
            for i, r in enumerate(sheet.iter_rows(values_only=True)):
                # Skip empty lines
                if not any(r):
                    continue
                if not headers:
                    headers = [str(h).strip() if h is not None else "" for h in r]
                else:
                    rows.append([str(v) if v is not None else "" for v in r])
                    if len(rows) >= 20:
                        break
                        
        return {"headers": headers, "preview_rows": rows}

    @classmethod
    async def process_import_batch(
        cls,
        db: AsyncSession,
        file_bytes: bytes,
        filename: str,
        column_mapping: Dict[str, str],
        mode: str,  # "candidate_only", "lead_only", "lead_candidate"
        upload_user: str,
        tag: Optional[str] = None
    ) -> Dict[str, Any]:
        """Parses entire sheet and imports records following selected mode and CRM duplicate check rules."""
        # Check mapping contains basic fields
        # Note: mapping is: { "full_name": "Excel Column Name", "email": "Excel Column Name" }
        inverse_mapping = {v: k for k, v in column_mapping.items() if v}
        
        rows = []
        if filename.endswith(".csv"):
            try:
                content = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                content = file_bytes.decode("latin-1")
            reader = csv.DictReader(io.StringIO(content))
            for r in reader:
                rows.append({inverse_mapping.get(k.strip(), k.strip()): v for k, v in r.items()})
        else:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet = wb.active
            headers = []
            for r in sheet.iter_rows(values_only=True):
                if not any(r):
                    continue
                if not headers:
                    headers = [str(h).strip() if h is not None else "" for h in r]
                else:
                    row_data = {}
                    for col_idx, val in enumerate(r):
                        if col_idx < len(headers):
                            h_name = headers[col_idx]
                            mapped_field = inverse_mapping.get(h_name)
                            if mapped_field:
                                row_data[mapped_field] = str(val) if val is not None else ""
                    if row_data:
                        rows.append(row_data)

        # Batch creation stats
        batch = CandidateImportBatch(
            id=str(uuid.uuid4()),
            file_name=filename,
            uploaded_by=upload_user,
            total_rows=len(rows),
            new_records=0,
            updated_records=0,
            duplicate_records=0,
            failed_records=0,
            created_at=datetime.utcnow()
        )
        db.add(batch)
        await db.flush()

        for idx, row in enumerate(rows):
            # Read mapped values
            email = row.get("email", "").strip().lower()
            phone = row.get("phone", "").strip()
            name = row.get("full_name", "").strip()
            
            # Normalize phone (keep primary number if multiple are specified)
            if phone:
                for sep in (",", "/", ";", " "):
                    if sep in phone:
                        phone = phone.split(sep)[0].strip()
                        break
            
            # Basic validation
            if not email and not phone:
                batch.failed_records += 1
                continue

            # Format/clean fields
            whatsapp = row.get("whatsapp_number", "").strip()
            if whatsapp:
                for sep in (",", "/", ";", " "):
                    if sep in whatsapp:
                        whatsapp = whatsapp.split(sep)[0].strip()
                        break
            address = row.get("address", "").strip()
            program_type = row.get("program_type", "").strip()
            emergency = row.get("emergency_contact", "").strip()
            qualification = row.get("qualification", "").strip()
            blood = row.get("blood_group", "").strip()
            course = row.get("course_applied", "").strip()
            mode_learning = row.get("mode_of_learning", "").strip()
            college = row.get("college_name", "").strip()
            dob_str = row.get("date_of_birth", "").strip()
            gender = row.get("gender", "").strip()
            reference = row.get("reference_details", "").strip()
            languages = row.get("languages_known", "").strip()
            parent = row.get("parent_guardian_name", "").strip()
            parent_occ = row.get("parent_guardian_occupation", "").strip()
            aadhaar = row.get("aadhaar_number", "").strip()
            remarks = row.get("remarks", "").strip()
            
            dob = None
            if dob_str:
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
                    try:
                        dob = datetime.strptime(dob_str.split()[0], fmt)
                        break
                    except Exception:
                        continue

            # -------------------------------------------------------------
            # Mode processing
            # -------------------------------------------------------------
            
            # MODE: LEAD ONLY or LEAD + CANDIDATE
            lead_id = None
            if mode in ("lead_only", "lead_candidate"):
                # Check for existing lead using standard CRM rules
                lead_stmt = select(Lead).where(
                    and_(
                        Lead.is_deleted == False,
                        or_(
                            Lead.email.ilike(email) if email else False,
                            Lead.phone == phone if phone else False
                        )
                    )
                )
                lead_res = await db.execute(lead_stmt)
                existing_lead = lead_res.scalars().first()
                
                if existing_lead:
                    # Duplicate Lead detected
                    existing_lead.duplicate_hits += 1
                    existing_lead.interaction_count += 1
                    existing_lead.last_interaction_at = datetime.utcnow()
                    existing_lead.latest_source = f"Excel Import - {batch.file_name}"
                    
                    # Merge course
                    if course:
                        merged = existing_lead.merged_courses or []
                        if course not in merged:
                            merged.append(course)
                        existing_lead.merged_courses = merged
                    
                    # Update empty lead fields
                    if name and not existing_lead.name:
                        existing_lead.name = name
                    if email and not existing_lead.email:
                        existing_lead.email = email
                    
                    # Log timeline and interaction
                    it = LeadInteraction(
                        id=str(uuid.uuid4()),
                        lead_id=existing_lead.id,
                        interaction_type="Import Duplicate",
                        source=f"Excel Import",
                        course=course or existing_lead.interested_course,
                        notes=f"Duplicate lead detected during batch import (Batch tag: {tag or 'None'}).",
                        created_at=datetime.utcnow()
                    )
                    db.add(it)
                    
                    lt = LeadTimelineEventModel(
                        id=str(uuid.uuid4()),
                        lead_id=existing_lead.id,
                        event_type="Import Updated",
                        description=f"Lead interaction merged from batch import '{filename}'.",
                        created_by=upload_user,
                        created_at=datetime.utcnow()
                    )
                    db.add(lt)
                    lead_id = existing_lead.id
                    batch.duplicate_records += 1
                else:
                    # Create new Lead
                    new_lead = Lead(
                        id=str(uuid.uuid4()),
                        name=name or "Unknown Candidate",
                        email=email or "",
                        phone=phone or "",
                        interested_course=course or "",
                        source=f"Excel Import - {batch.file_name}",
                        status="Pending",
                        duplicate_hits=0,
                        interaction_count=1,
                        last_interaction_at=datetime.utcnow(),
                        first_source=f"Excel Import",
                        latest_source=f"Excel Import",
                        created_at=datetime.utcnow()
                    )
                    db.add(new_lead)
                    await db.flush()
                    
                    lt = LeadTimelineEventModel(
                        id=str(uuid.uuid4()),
                        lead_id=new_lead.id,
                        event_type="Import Created",
                        description=f"Lead created via batch import '{filename}'.",
                        created_by=upload_user,
                        created_at=datetime.utcnow()
                    )
                    db.add(lt)
                    lead_id = new_lead.id
                    batch.new_records += 1

            # MODE: CANDIDATE ONLY or LEAD + CANDIDATE
            if mode in ("candidate_only", "lead_candidate"):
                # Duplicate check inside Candidate table
                cand_stmt = select(CandidateApplication).where(
                    or_(
                        CandidateApplication.email.ilike(email) if email else False,
                        CandidateApplication.phone == phone if phone else False
                    ),
                    CandidateApplication.is_deleted == False
                )
                cand_res = await db.execute(cand_stmt)
                existing_candidate = cand_res.scalars().first()
                
                if existing_candidate:
                    # Update missing fields in candidate
                    if name and not existing_candidate.full_name:
                        existing_candidate.full_name = name
                    if whatsapp and not existing_candidate.whatsapp_number:
                        existing_candidate.whatsapp_number = whatsapp
                    if address and not existing_candidate.address:
                        existing_candidate.address = address
                    if emergency and not existing_candidate.emergency_contact:
                        existing_candidate.emergency_contact = emergency
                    if qualification and not existing_candidate.qualification:
                        existing_candidate.qualification = qualification
                    if blood and not existing_candidate.blood_group:
                        existing_candidate.blood_group = blood
                    if course and not existing_candidate.course_applied:
                        existing_candidate.course_applied = course
                    if mode_learning and not existing_candidate.mode_of_learning:
                        existing_candidate.mode_of_learning = mode_learning
                    if college and not existing_candidate.college_name:
                        existing_candidate.college_name = college
                    if dob and not existing_candidate.date_of_birth:
                        existing_candidate.date_of_birth = dob
                    if gender and not existing_candidate.gender:
                        existing_candidate.gender = gender
                    if reference and not existing_candidate.reference_details:
                        existing_candidate.reference_details = reference
                    if languages and not existing_candidate.languages_known:
                        existing_candidate.languages_known = languages
                    if parent and not existing_candidate.parent_guardian_name:
                        existing_candidate.parent_guardian_name = parent
                    if parent_occ and not existing_candidate.parent_guardian_occupation:
                        existing_candidate.parent_guardian_occupation = parent_occ
                    if program_type and not existing_candidate.program_type:
                        existing_candidate.program_type = program_type
                    if aadhaar and not existing_candidate.aadhaar_number_encrypted:
                        existing_candidate.aadhaar_number_encrypted = encrypt_aadhaar(aadhaar)
                    if remarks:
                        await cls.add_candidate_note(
                            db,
                            existing_candidate.id,
                            f"[Import update remarks]: {remarks}",
                            created_by=upload_user
                        )
                    
                    existing_candidate.updated_at = datetime.utcnow()
                    
                    # Track import batch
                    existing_candidate.import_batch_id = batch.id
                    existing_candidate.import_tag = tag
                    
                    # Add notes and timeline
                    await cls.add_timeline_event(
                        db,
                        existing_candidate.id,
                        "Import Merged",
                        f"Candidate details updated during Excel batch import '{filename}' (Mode: {mode}).",
                        upload_user
                    )
                    
                    if mode == "candidate_only":
                        batch.duplicate_records += 1
                    else:
                        batch.updated_records += 1
                else:
                    # Create new Candidate Application
                    app_num = await cls.generate_application_number(db)
                    
                    new_candidate = CandidateApplication(
                        id=str(uuid.uuid4()),
                        lead_id=lead_id,
                        application_number=app_num,
                        full_name=name or "Unknown Candidate",
                        email=email or "",
                        phone=phone or "",
                        whatsapp_number=whatsapp,
                        address=address,
                        emergency_contact=emergency,
                        qualification=qualification,
                        blood_group=blood,
                        course_applied=course,
                        mode_of_learning=mode_learning or "Offline",
                        college_name=college,
                        date_of_birth=dob,
                        gender=gender,
                        reference_details=reference,
                        languages_known=languages,
                        parent_guardian_name=parent,
                        parent_guardian_occupation=parent_occ,
                        aadhaar_number_encrypted=encrypt_aadhaar(aadhaar) if aadhaar else None,
                        program_type=program_type if program_type else None,
                        application_status="Submitted",
                        document_status="Missing Documents",  # Since Excel import has no docs initially
                        candidate_source="Excel Import",
                        import_batch_id=batch.id,
                        import_tag=tag,
                        created_at=datetime.utcnow(),
                        updated_at=datetime.utcnow()
                    )
                    
                    db.add(new_candidate)
                    await db.flush()

                    if remarks:
                        await cls.add_candidate_note(
                            db,
                            new_candidate.id,
                            f"[Import remarks]: {remarks}",
                            created_by=upload_user
                        )
                    
                    await cls.add_timeline_event(
                        db,
                        new_candidate.id,
                        "Submitted",
                        f"Candidate application created during Excel batch import '{filename}' (Mode: {mode}).",
                        upload_user
                    )
                    
                    if mode == "candidate_only":
                        batch.new_records += 1

        await db.commit()
        return {
            "batch_id": batch.id,
            "total_rows": batch.total_rows,
            "new_records": batch.new_records,
            "updated_records": batch.updated_records,
            "duplicate_records": batch.duplicate_records,
            "failed_records": batch.failed_records
        }

    @classmethod
    async def execute_import_in_background(
        cls,
        file_bytes: bytes,
        filename: str,
        column_mapping: Dict[str, str],
        mode: str,
        upload_user: str,
        tag: Optional[str] = None
    ) -> None:
        """Runs process_import_batch asynchronously in a background task with a new database session."""
        from app.core.database import AsyncSessionLocal
        async with AsyncSessionLocal() as db:
            try:
                await cls.process_import_batch(
                    db=db,
                    file_bytes=file_bytes,
                    filename=filename,
                    column_mapping=column_mapping,
                    mode=mode,
                    upload_user=upload_user,
                    tag=tag
                )
            except Exception as e:
                logger.error(f"Error executing background import for file {filename}: {e}", exc_info=True)

    @classmethod
    async def generate_and_upload_receipt(cls, db: AsyncSession, payment_id: str) -> None:
        """
        Generates a PDF receipt for a payment, uploads it to Supabase Storage,
        and updates the payment record with receipt_number and receipt_url.
        """
        from app.models.candidate_payment import CandidatePayment
        
        # 1. Fetch payment and candidate details
        stmt = select(CandidatePayment).where(CandidatePayment.id == payment_id).options(
            selectinload(CandidatePayment.candidate).selectinload(CandidateApplication.payments)
        )
        res = await db.execute(stmt)
        payment = res.scalar_one_or_none()
        if not payment:
            logger.error(f"Payment with ID {payment_id} not found for receipt generation.")
            return

        candidate = payment.candidate
        
        # 2. Generate unique receipt number if not present
        if not payment.receipt_number:
            payment.receipt_number = f"REC-{payment.id[:8].upper()}"
            
        # 3. Generate PDF
        try:
            import io
            import os
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.lib.colors import HexColor
            from reportlab.pdfgen import canvas
            from reportlab.lib.utils import ImageReader
            
            pdf_buffer = io.BytesIO()
            c = canvas.Canvas(pdf_buffer, pagesize=A4)
            width, height = A4
            margin = 22 * mm
            content_w = width - 2 * margin
            
            # Branded design system colors
            NAVY = HexColor("#16263F")
            TEAL = HexColor("#1F9C9C")
            DARK_TEXT = HexColor("#262B33")
            GREY_TEXT = HexColor("#6B7280")
            HAIRLINE = HexColor("#D8DEE6")
            
            # Top accent bar
            c.setFillColor(NAVY)
            c.rect(0, height - 4 * mm, width, 4 * mm, fill=1, stroke=0)
            c.setFillColor(TEAL)
            c.rect(0, height - 4 * mm, width * 0.32, 4 * mm, fill=1, stroke=0)
            
            # Header (centered logo & text)
            top = height - 18 * mm
            logo_size = 18 * mm
            
            logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "AgenticX-removebg-preview.png")
            try:
                logo_img = ImageReader(logo_path)
                c.drawImage(
                    logo_img,
                    width / 2 - logo_size / 2,
                    top - logo_size,
                    width=logo_size,
                    height=logo_size,
                    mask="auto",
                    preserveAspectRatio=True,
                )
            except Exception:
                c.setFillColor(NAVY)
                c.roundRect(width / 2 - logo_size / 2, top - logo_size, logo_size, logo_size, 3 * mm, fill=1, stroke=0)
                c.setFillColor(HexColor("#FFFFFF"))
                c.setFont("Helvetica-Bold", 8)
                c.drawCentredString(width / 2, top - logo_size / 2 - 2, "AgenticX")
                
            name_y = top - logo_size - 6 * mm
            c.setFillColor(NAVY)
            c.setFont("Helvetica-Bold", 20)
            c.drawCentredString(width / 2, name_y, "AgenticX Knowledge Solutions")
            
            c.setFillColor(GREY_TEXT)
            c.setFont("Helvetica", 9)
            c.drawCentredString(width / 2, name_y - 5 * mm, "3rd Floor, Raj Plaza, Town Limit, Kollam, Kerala")
            c.drawCentredString(width / 2, name_y - 9 * mm, "www.agenticx.co.in  |  anju.muraleedharan@agenticx.co.in  |  +91 94965 52094")
            
            rule_y = name_y - 14 * mm
            c.setStrokeColor(TEAL)
            c.setLineWidth(1.2)
            c.line(margin, rule_y, width - margin, rule_y)
            
            # Receipt Title & Meta
            title_y = rule_y - 10 * mm
            c.setFillColor(NAVY)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(margin, title_y, "PAYMENT RECEIPT")
            
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(DARK_TEXT)
            c.drawRightString(width - margin, title_y, f"Receipt No: {payment.receipt_number}")
            c.setFont("Helvetica", 10)
            c.setFillColor(GREY_TEXT)
            pay_date_str = payment.payment_date.strftime('%d-%b-%Y') if payment.payment_date else datetime.now().strftime('%d-%b-%Y')
            c.drawRightString(width - margin, title_y - 5 * mm, f"Date: {pay_date_str}")
            
            # Details Section Box
            box_top = title_y - 12 * mm
            box_h = 88 * mm
            c.setStrokeColor(HAIRLINE)
            c.setFillColor(HexColor("#F9FAFB"))
            c.roundRect(margin, box_top - box_h, content_w, box_h, 3 * mm, fill=1, stroke=1)
            
            # Candidate Info (Left Column)
            c.setFillColor(DARK_TEXT)
            c.setFont("Helvetica-Bold", 11)
            col1_x = margin + 6 * mm
            col2_x = margin + content_w / 2 + 4 * mm
            
            y = box_top - 8 * mm
            c.drawString(col1_x, y, "Candidate Details")
            c.setFont("Helvetica", 10)
            c.setFillColor(DARK_TEXT)
            # Resolve CAF/Application number dynamically without hardcoding
            caf_number_val = None
            for field in ["application_number", "caf_number", "application_id", "id"]:
                if hasattr(candidate, field):
                    val = getattr(candidate, field)
                    if val:
                        caf_number_val = val
                        break
            if not caf_number_val:
                caf_number_val = getattr(candidate, "id", "N/A")

            # Resolve Program/Course name dynamically without hardcoding
            course_name_val = None
            for field in ["course_applied", "course_name", "program_name"]:
                if hasattr(candidate, field):
                    val = getattr(candidate, field)
                    if val:
                        course_name_val = val
                        break
            if not course_name_val and hasattr(candidate, "program") and candidate.program:
                course_name_val = getattr(candidate.program, "name", None)
            if not course_name_val:
                course_name_val = "N/A"

            c.drawString(col1_x, y - 6 * mm, f"Name: {candidate.full_name}")
            c.drawString(col1_x, y - 12 * mm, f"Email: {candidate.email}")
            c.drawString(col1_x, y - 18 * mm, f"Phone: {candidate.phone}")
            c.drawString(col1_x, y - 24 * mm, f"CAF Number: {caf_number_val}")
            
            c.setFillColor(DARK_TEXT)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(col1_x, y - 36 * mm, "Program Details")
            c.setFont("Helvetica", 10)
            c.drawString(col1_x, y - 42 * mm, f"Program: {course_name_val}")
            c.drawString(col1_x, y - 48 * mm, f"Type: {candidate.program_type or 'N/A'}")
            
            # Payment Info (Right Column)
            c.setFillColor(DARK_TEXT)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(col2_x, y, "Payment Details")
            c.setFont("Helvetica", 10)
            c.drawString(col2_x, y - 6 * mm, f"Payment Type: {payment.payment_type}")
            c.drawString(col2_x, y - 12 * mm, f"Method: {payment.payment_method}")
            txn_id = payment.transaction_id or "N/A"
            if len(txn_id) > 28:
                c.drawString(col2_x, y - 18 * mm, "Transaction ID:")
                c.setFont("Helvetica", 8)
                c.drawString(col2_x, y - 22 * mm, txn_id)
                c.setFont("Helvetica", 10)
                y_offset = 4 * mm
            else:
                c.drawString(col2_x, y - 18 * mm, f"Transaction ID: {txn_id}")
                y_offset = 0.0
                
            c.drawString(col2_x, y - 24 * mm - y_offset, f"Status: {payment.status}")
            
            # Financial Overview (Centralized source of truth)
            financials = cls.calculate_financials(candidate, current_payment=payment)

            c.setFillColor(DARK_TEXT)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(col2_x, y - 32 * mm - y_offset, "Financial Overview")
            c.setFont("Helvetica", 9)
            c.drawString(col2_x, y - 37 * mm - y_offset, f"Course Fee: INR {financials['final_payable_amount']:.2f}")
            adm_status_label = "Paid" if financials["admission_fee_paid"] else "Pending"
            c.drawString(col2_x, y - 42 * mm - y_offset, f"Admission Fee: INR {financials['admission_fee_amount']:.2f} ({adm_status_label})")
            c.drawString(col2_x, y - 47 * mm - y_offset, f"Course Fee Paid: INR {financials['course_paid']:.2f}")
            c.drawString(col2_x, y - 52 * mm - y_offset, f"Remaining Course Fee: INR {financials['remaining_course_balance']:.2f}")
            c.drawString(col2_x, y - 57 * mm - y_offset, f"Total Collected: INR {financials['total_collected']:.2f}")
            
            c.setFont("Helvetica-Bold", 11)
            c.setFillColor(HexColor("#10B981"))
            c.drawString(col2_x, y - 66 * mm - y_offset, f"Amount Paid (This Receipt): INR {payment.amount:.2f}")
            
            # Footer
            sig_y = box_top - box_h - 25 * mm
            c.setStrokeColor(HAIRLINE)
            c.setLineWidth(0.5)
            c.line(margin, sig_y + 10 * mm, width - margin, sig_y + 10 * mm)
            
            c.setFont("Helvetica", 8)
            c.setFillColor(GREY_TEXT)
            c.drawString(margin, sig_y, "This is a computer-generated document. No signature required.")
            c.drawRightString(width - margin, sig_y, "Thank you for choosing AgenticX.")
            
            c.showPage()
            c.save()
            pdf_bytes = pdf_buffer.getvalue()
            
            # Upload to storage
            uploader = UploadService()
            url = await uploader.upload_file(
                file_content=pdf_bytes,
                folder="receipts",
                original_filename=f"receipt_{payment.receipt_number}.pdf",
                mime_type="application/pdf"
            )
            payment.receipt_url = url
            await db.commit()
            logger.info(f"Generated and uploaded receipt {payment.receipt_number} to {url}")
            
        except Exception as e:
            logger.error(f"Error generating receipt PDF for payment {payment_id}: {e}", exc_info=True)

