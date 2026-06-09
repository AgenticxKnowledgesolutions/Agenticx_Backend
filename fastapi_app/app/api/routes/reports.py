import csv
import io
from datetime import datetime, timedelta
from typing import Optional, List, Any
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from app.core.database import get_db
from app.models.lead import Lead
from app.schemas.lead import LeadResponse
from app.deps import require_admin
from app.models.user import User

router = APIRouter(prefix="/reports", tags=["Reports"])

def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    if not date_str:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except ValueError:
        return None


async def get_filtered_leads_query(date_from: Optional[str], date_to: Optional[str]):
    stmt = select(Lead).where(Lead.is_deleted == False)
    
    if date_from:
        d_from = parse_date(date_from)
        if d_from:
            stmt = stmt.where(Lead.created_at >= d_from)
            
    if date_to:
        d_to = parse_date(date_to)
        if d_to:
            # Add 1 day if it is just a date like YYYY-MM-DD
            if len(date_to) <= 10:
                d_to = d_to + timedelta(days=1) - timedelta(seconds=1)
            stmt = stmt.where(Lead.created_at <= d_to)
            
    return stmt.order_by(Lead.created_at.desc())


@router.get("/leads", response_model=List[LeadResponse])
async def get_report_leads(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin)
):
    stmt = await get_filtered_leads_query(date_from, date_to)
    result = await db.execute(stmt)
    return result.scalars().all()


@router.get("/leads/export")
async def export_report_leads(
    format: str = Query("csv"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin)
):
    stmt = await get_filtered_leads_query(date_from, date_to)
    result = await db.execute(stmt)
    leads = result.scalars().all()

    if format.lower() == "csv":
        return generate_csv_export(leads)
    elif format.lower() == "xlsx":
        return generate_xlsx_export(leads)
    elif format.lower() == "pdf":
        return generate_pdf_export(leads, date_from, date_to)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported format '{format}'. Use csv, xlsx, or pdf."
        )


def generate_csv_export(leads: List[Lead]):
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    writer.writerow([
        "Lead ID", "Student Name", "Email", "Phone", "Interested Course", 
        "Source Page", "Source", "Status", "Admin Notes", "Follow-up Notes", 
        "Last Contacted At", "Next Follow-up Date", "Created Date"
    ])
    
    # Rows
    for lead in leads:
        writer.writerow([
            lead.id,
            lead.name,
            lead.email,
            lead.phone or "",
            lead.interested_course or "",
            lead.source_page or "",
            lead.source or "Website",
            lead.status,
            lead.admin_notes or "",
            lead.followup_notes or "",
            lead.last_contacted_at.isoformat() if lead.last_contacted_at else "",
            lead.next_followup_date.isoformat() if lead.next_followup_date else "",
            lead.created_at.isoformat() if lead.created_at else ""
        ])
        
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=leads_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def generate_xlsx_export(leads: List[Lead]):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Report"
    
    # Style definitions
    font_title = Font(name="Arial", size=16, bold=True, color="001943")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_cell = Font(name="Arial", size=10)
    
    fill_header = PatternFill(start_color="001943", end_color="001943", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Title Block
    ws.merge_cells("A1:M1")
    ws["A1"] = "AgenticX Leads Operational Report"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Spacing
    ws.row_dimensions[2].height = 15
    
    # Headers
    headers = [
        "Lead ID", "Student Name", "Email", "Phone", "Interested Course", 
        "Source Page", "Source", "Status", "Admin Notes", "Follow-up Notes", 
        "Last Contacted At", "Next Follow-up Date", "Created Date"
    ]
    ws.append(headers)
    ws.row_dimensions[3].height = 26
    
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Rows
    for r_idx, lead in enumerate(leads, 4):
        ws.append([
            lead.id,
            lead.name,
            lead.email,
            lead.phone or "",
            lead.interested_course or "",
            lead.source_page or "",
            lead.source or "Website",
            lead.status,
            lead.admin_notes or "",
            lead.followup_notes or "",
            lead.last_contacted_at.strftime("%Y-%m-%d %H:%M") if lead.last_contacted_at else "",
            lead.next_followup_date.strftime("%Y-%m-%d %H:%M") if lead.next_followup_date else "",
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else ""
        ])
        ws.row_dimensions[r_idx].height = 20
        
        # Zebra striping and borders
        use_zebra = (r_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = font_cell
            cell.border = border_thin
            if use_zebra:
                cell.fill = fill_zebra
                
    from openpyxl.utils import get_column_letter
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            # Skip title row
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to BytesIO
    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    
    return StreamingResponse(
        out_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=leads_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
    )


def generate_pdf_export(leads: List[Lead], date_from: Optional[str], date_to: Optional[str]):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    out_stream = io.BytesIO()
    doc = SimpleDocTemplate(
        out_stream, 
        pagesize=letter, 
        rightMargin=36, 
        leftMargin=36, 
        topMargin=36, 
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    style_title = ParagraphStyle(
        name="TitleStyle",
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#001943"),
        spaceAfter=15,
        alignment=1 # Center
    )
    
    style_meta = ParagraphStyle(
        name="MetaStyle",
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=20,
        alignment=1 # Center
    )

    style_section = ParagraphStyle(
        name="SectionStyle",
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.HexColor("#001943"),
        spaceBefore=12,
        spaceAfter=6
    )

    style_body_bold = ParagraphStyle(
        name="BodyBoldStyle",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.HexColor("#001943")
    )
    
    style_cell = ParagraphStyle(
        name="CellTextStyle",
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#1e293b")
    )

    style_cell_header = ParagraphStyle(
        name="CellHeaderStyle",
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.white
    )

    story = []
    
    # 1. Title Page Header
    story.append(Paragraph("AgenticX Lead Operations & Intelligence Report", style_title))
    date_str = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if date_from or date_to:
        date_str += f" | Filter Range: {date_from or 'Start'} to {date_to or 'Now'}"
    story.append(Paragraph(date_str, style_meta))
    
    # 2. Compute KPI summary figures
    total_count = len(leads)
    contacted_count = sum(1 for l in leads if l.status.lower() in ["contacted", "demo booked", "enrolled", "converted"])
    enrolled_count = sum(1 for l in leads if l.status.lower() in ["enrolled", "converted"])
    pending_count = sum(1 for l in leads if l.status.lower() == "pending")
    conversion_rate = int((enrolled_count / total_count) * 100) if total_count > 0 else 0
    
    # KPI Box table
    kpi_data = [
        [
            Paragraph("Total Leads", style_body_bold),
            Paragraph("Pending Leads", style_body_bold),
            Paragraph("Contacted Leads", style_body_bold),
            Paragraph("Enrolled Leads", style_body_bold),
            Paragraph("Conversion Rate", style_body_bold)
        ],
        [
            Paragraph(str(total_count), style_cell),
            Paragraph(str(pending_count), style_cell),
            Paragraph(str(contacted_count), style_cell),
            Paragraph(str(enrolled_count), style_cell),
            Paragraph(f"{conversion_rate}%", style_cell)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[108, 108, 108, 108, 108])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f8fafc")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor("#cbd5e1")),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(Paragraph("Executive Performance Summary", style_section))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # 3. Status and Course distributions side-by-side
    status_counts = {}
    course_counts = {}
    for l in leads:
        status_counts[l.status] = status_counts.get(l.status, 0) + 1
        course_name = l.interested_course or "General Inquiry"
        course_counts[course_name] = course_counts.get(course_name, 0) + 1
        
    status_table_data = [[Paragraph("Lead Status", style_body_bold), Paragraph("Count", style_body_bold)]]
    for s_name, count in status_counts.items():
        status_table_data.append([Paragraph(s_name or "Pending", style_cell), Paragraph(str(count), style_cell)])
        
    course_table_data = [[Paragraph("Course Name", style_body_bold), Paragraph("Count", style_body_bold)]]
    for c_name, count in sorted(course_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
        course_table_data.append([Paragraph(c_name, style_cell), Paragraph(str(count), style_cell)])
        
    # Side-by-side Table containers
    t_status = Table(status_table_data, colWidths=[150, 80])
    t_status.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    t_course = Table(course_table_data, colWidths=[170, 80])
    t_course.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#f1f5f9")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    distribution_layout_data = [
        [Paragraph("Lead Status Distribution", style_section), Paragraph("Top Course Interests", style_section)],
        [t_status, t_course]
    ]
    
    distribution_table = Table(distribution_layout_data, colWidths=[260, 280])
    distribution_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    
    story.append(distribution_table)
    story.append(Spacer(1, 15))
    
    # 4. Leads Detail Table
    story.append(Paragraph("Detailed Lead Log", style_section))
    
    leads_table_data = [[
        Paragraph("Name", style_cell_header),
        Paragraph("Email", style_cell_header),
        Paragraph("Phone", style_cell_header),
        Paragraph("Course", style_cell_header),
        Paragraph("Source", style_cell_header),
        Paragraph("Status", style_cell_header),
        Paragraph("Date", style_cell_header)
    ]]
    
    for l in leads:
        leads_table_data.append([
            Paragraph(l.name or "N/A", style_cell),
            Paragraph(l.email or "N/A", style_cell),
            Paragraph(l.phone or "N/A", style_cell),
            Paragraph(l.interested_course or "General", style_cell),
            Paragraph(l.source or "Website", style_cell),
            Paragraph(l.status or "Pending", style_cell),
            Paragraph(l.created_at.strftime("%Y-%m-%d") if l.created_at else "N/A", style_cell)
        ])
        
    # Standard margins let us use a width of 540 max (letter = 612 wide. 612 - 72 = 540)
    col_widths = [90, 110, 75, 95, 60, 60, 50]
    
    leads_table = Table(leads_table_data, colWidths=col_widths, repeatRows=1)
    leads_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#001943")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#cbd5e1")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    
    story.append(leads_table)
    
    doc.build(story)
    out_stream.seek(0)
    
    return StreamingResponse(
        out_stream,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=leads_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"}
    )
